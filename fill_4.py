"""
Example script for testing the Forest theme

Author: rdbende
License: MIT license
Source: https://github.com/rdbende/ttk-widget-factory
"""

import tkinter as tk
from tkinter import ttk
from tkinter import *
from openpyxl import *
import openpyxl

# Update the listbox
def update(data):
    # Clear the listbox
    my_list.delete(0,END)

    # Add toppings to listbox
    for item in data:
        my_list.insert(END,item)

def fillout(e):
    my_entry.delete(0,END)

    my_entry.insert(0,my_list.get(ACTIVE))

def check(e):

    typed = my_entry.get()

    if typed == '':
        data = toppings
    else:
        data = []
        for item in toppings:
            if typed.lower() in item.lower():
                data.append(item)

    update(data)

root = tk.Tk()
root.title("Forest")
root.option_add("*tearOff", False) # This is always a good idea

g = tk.DoubleVar(value=75.0)

# Make the app responsive
root.columnconfigure(index=0, weight=1)
root.columnconfigure(index=1, weight=1)
root.columnconfigure(index=2, weight=1)
root.rowconfigure(index=0, weight=1)
root.rowconfigure(index=1, weight=1)
root.rowconfigure(index=2, weight=1)

# Create a style
style = ttk.Style(root)

# Import the tcl file
root.tk.call("source", "forest-dark.tcl")

# Set the theme with the theme_use method
style.theme_use("forest-dark")



# Panedwindow
paned = ttk.PanedWindow(root)
paned.grid(row=0, column=0, pady=(25, 5), sticky="nsew", rowspan=3)

# Pane #2
pane_2 = ttk.Frame(paned)
paned.add(pane_2, weight=3)

# Notebook
notebook = ttk.Notebook(pane_2)

# Tab #1
tab_1 = ttk.Frame(notebook)
tab_1.columnconfigure(index=0, weight=1)
tab_1.columnconfigure(index=1, weight=1)
tab_1.rowconfigure(index=0, weight=1)
tab_1.rowconfigure(index=1, weight=1)
notebook.add(tab_1, text="Productos")

# Fill_3


wb = load_workbook('Lista_Reyes.xlsx')
ws = wb.active

m_row = ws.max_row
print(ws)
print(m_row)
toppings = []
for i in range(1,m_row+1):
    cell_obj = ws.cell(row=i, column=1)
    print(cell_obj.value)
    toppings.append(cell_obj.value)
    

my_label = Label(tab_1, text="Start Typing...", font=("Helvetica",14),fg="grey")
my_label.grid(row=0, column=0, padx=(20, 10), pady=(20, 0), sticky="ew")

my_entry = Entry(tab_1,font=("Helvetica",20))
my_entry.grid(row=1, column=0, padx=(10, 20), pady=(20, 0), sticky="ew")

my_list = Listbox(tab_1,width=50)
my_list.grid(row=2, column=0, padx=(10, 20), pady=(20, 0), sticky="ew")

##toppings = ["Peperoni","Peppers","Mushrooms","Cheese","Onions","Ham","Taco"]


update(toppings)

my_list.bind("<<ListboxSelect>>",fillout)

my_entry.bind("<KeyRelease>",check)

### Scale
##scale = ttk.Scale(tab_1, from_=100, to=0, variable=g, command=lambda event: g.set(scale.get()))
##scale.grid(row=0, column=0, padx=(20, 10), pady=(20, 0), sticky="ew")
##
### Progressbar
##progress = ttk.Progressbar(tab_1, value=0, variable=g, mode="determinate")
##progress.grid(row=0, column=1, padx=(10, 20), pady=(20, 0), sticky="ew")
##
### Label
##label = ttk.Label(tab_1, text="Forest ttk theme", justify="center")
##label.grid(row=1, column=0, pady=10, columnspan=2)

# Tab #2
tab_2 = ttk.Frame(notebook)
notebook.add(tab_2, text="Tab 2")

# Tab #3
tab_3 = ttk.Frame(notebook)
notebook.add(tab_3, text="Tab 3")

notebook.pack(expand=True, fill="both", padx=5, pady=5)

# Sizegrip
sizegrip = ttk.Sizegrip(root)
sizegrip.grid(row=100, column=100, padx=(0, 5), pady=(0, 5))

# Center the window, and set minsize
root.update()
root.minsize(root.winfo_width(), root.winfo_height())
x_cordinate = int((root.winfo_screenwidth()/2) - (root.winfo_width()/2))
y_cordinate = int((root.winfo_screenheight()/2) - (root.winfo_height()/2))
root.geometry("+{}+{}".format(x_cordinate, y_cordinate))

# Start the main loop
root.mainloop()
