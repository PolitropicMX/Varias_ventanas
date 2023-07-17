"""
Example script for testing the Forest theme

Author: rdbende
License: MIT license
Source: https://github.com/rdbende/ttk-widget-factory
"""


import tkinter as tk
from tkinter import ttk
from tkinter import *
from openpyxl import *
import openpyxl

# Update the listbox
def update(data):
    # Clear the listbox
    my_list.delete(0,END)

    # Add toppings to listbox
    for item in data:
        my_list.insert(END,item)

def fillout(e):
    my_entry.delete(0,END)

    my_entry.insert(0,my_list.get(ACTIVE))

def check(e):

    typed = my_entry.get()

    if typed == '':
        data = toppings
    else:
        data = []
        for item in toppings:
            if typed.lower() in item.lower():
                data.append(item)

    update(data)

root = tk.Tk()
root.title("Forest")
root.option_add("*tearOff", False) # This is always a good idea

# Make the app responsive
root.columnconfigure(index=0, weight=1)
root.columnconfigure(index=1, weight=1)
root.columnconfigure(index=2, weight=1)
root.rowconfigure(index=0, weight=1)
root.rowconfigure(index=1, weight=1)
root.rowconfigure(index=2, weight=1)

# Create a style
style = ttk.Style(root)

# Import the tcl file
root.tk.call("source", "forest-dark.tcl")

# Set the theme with the theme_use method
style.theme_use("forest-dark")

########################################################################

wb = load_workbook('Lista_Reyes.xlsx')
ws = wb.active

m_row = ws.max_row
print(ws)
print(m_row)
toppings = []
for i in range(1,m_row+1):
    cell_obj = ws.cell(row=i, column=1)
    print(cell_obj.value)
    toppings.append(cell_obj.value)
    

my_label = Label(root, text="Start Typing...", font=("Helvetica",14),fg="grey")
my_label.pack(pady=20)

my_entry = Entry(root,font=("Helvetica",20))
my_entry.pack()

my_list = Listbox(root,width=50)
my_list.pack(pady=40)

##toppings = ["Peperoni","Peppers","Mushrooms","Cheese","Onions","Ham","Taco"]


update(toppings)

my_list.bind("<<ListboxSelect>>",fillout)

my_entry.bind("<KeyRelease>",check)


########################################################################

# Center the window, and set minsize
root.update()
root.minsize(root.winfo_width(), root.winfo_height())
x_cordinate = int((root.winfo_screenwidth()/2) - (root.winfo_width()/2))
y_cordinate = int((root.winfo_screenheight()/2) - (root.winfo_height()/2))
root.geometry("+{}+{}".format(x_cordinate, y_cordinate))

# Start the main loop
root.mainloop()
